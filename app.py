from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import json
import os
from datetime import datetime
from db_connection import get_connection
from modules.color_manager import ColorManager
from modules.bag_designer import BagDesigner
from modules.inventory_manager import InventoryManager
from modules.quotation_manager import QuotationManager
from modules.orders_manager import OrdersManager
from modules.backup_manager import BackupManager

app = Flask(__name__)

# ==================== FUNCIÓN HELPER PARA PDFs ====================
def dibujar_encabezado_pdf(c, width, height, titulo_documento="Documento"):
    """
    Dibuja el encabezado estándar en los PDFs con logo y título del sistema
    """
    from reportlab.lib.utils import ImageReader
    
    # Ruta del logo
    logo_path = os.path.join(app.static_folder, 'images', 'logo_.png')
    
    # Dibujar logo si existe - ESQUINA SUPERIOR IZQUIERDA
    if os.path.exists(logo_path):
        try:
            # Logo en la esquina superior izquierda (80x80 px)
            c.drawImage(logo_path, 40, height - 100, width=80, height=80, preserveAspectRatio=True, mask='auto')
        except:
            pass
    
    # Título del sistema - CENTRADO
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width / 2, height - 50, "Sistema Integral de Gestión")
    c.drawCentredString(width / 2, height - 68, "para Confeccionistas de Bolsas")
    
    # Marca ChromaBags - CENTRADO
    c.setFont("Helvetica-Bold", 20)
    c.drawCentredString(width / 2, height - 95, "CHROMABAGS")
    
    # Línea divisoria
    c.setStrokeColorRGB(0.7, 0.3, 0.5)  # Color morado/rosa
    c.setLineWidth(2)
    c.line(40, height - 110, width - 40, height - 110)
    
    # Título del documento - Alineado a la izquierda
    c.setFont("Helvetica-Bold", 16)
    c.setFillColorRGB(0, 0, 0)
    c.drawString(40, height - 140, titulo_documento)
    
    return height - 160  # Retorna la posición Y donde puede empezar el contenido

# Página principal
@app.route('/')
def index():
    return redirect(url_for('clientes'))

# ==================== CLIENTES ====================
@app.route('/clientes')
def clientes():
    conn = get_connection()
    clientes_data = []
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT nombre_cliente, telefono, correo, tipo_cliente, direccion, id_cliente,
                   rfc, razon_social, uso_cfdi, regimen_fiscal, correo_facturacion
            FROM clientes 
            ORDER BY id_cliente;
        """)
        clientes_data = cursor.fetchall()
        cursor.close()
        conn.close()
    return render_template('clientes.html', clientes=clientes_data)

@app.route('/agregar_cliente', methods=['POST'])
def agregar_cliente():
    # Datos básicos
    nombre = request.form['nombre'].upper()
    telefono = request.form['telefono']
    correo = request.form['correo']
    tipo = request.form['tipo']
    direccion = request.form['direccion'].upper()
    
    # Datos fiscales opcionales
    rfc = request.form.get('rfc', '').upper() or None
    razon_social = request.form.get('razon_social', '').upper() or None
    uso_cfdi = request.form.get('uso_cfdi') or None
    regimen_fiscal = request.form.get('regimen_fiscal') or None
    correo_facturacion = request.form.get('correo_facturacion') or None

    conn = get_connection()
    if conn:
        cursor = conn.cursor()
        
        # Verificar si la tabla tiene las columnas fiscales, si no, crearlas
        try:
            cursor.execute("""
                INSERT INTO clientes (
                    nombre_cliente, telefono, correo, tipo_cliente, direccion,
                    rfc, razon_social, uso_cfdi, regimen_fiscal, correo_facturacion
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
            """, (nombre, telefono, correo, tipo, direccion, 
                  rfc, razon_social, uso_cfdi, regimen_fiscal, correo_facturacion))
        except Exception as e:
            # Si las columnas no existen, agregarlas
            print(f"Error insertando cliente: {e}")
            cursor.execute("""
                INSERT INTO clientes (nombre_cliente, telefono, correo, tipo_cliente, direccion)
                VALUES (?, ?, ?, ?, ?);
            """, (nombre, telefono, correo, tipo, direccion))
        
        conn.commit()
        cursor.close()
        conn.close()
    return redirect(url_for('clientes'))

@app.route('/eliminar_cliente/<int:id>', methods=['GET', 'POST'])
def eliminar_cliente(id):
    conn = get_connection()
    if conn:
        cur = conn.cursor()
        try:
            # Verificar si el cliente tiene cotizaciones
            cur.execute("SELECT COUNT(*) FROM cotizaciones WHERE id_cliente = ?", (id,))
            tiene_cotizaciones = cur.fetchone()[0]
            
            # Verificar si el cliente tiene pedidos
            cur.execute("SELECT COUNT(*) FROM pedidos WHERE id_cliente = ?", (id,))
            tiene_pedidos = cur.fetchone()[0]
            
            if tiene_cotizaciones > 0 or tiene_pedidos > 0:
                print(f"⚠️ No se puede eliminar: cliente tiene {tiene_cotizaciones} cotizaciones y {tiene_pedidos} pedidos")
            else:
                cur.execute("DELETE FROM clientes WHERE id_cliente = ?;", (id,))
                conn.commit()
                print(f"✓ Cliente #{id} eliminado correctamente")
        except Exception as e:
            print(f"❌ Error al eliminar cliente: {e}")
            conn.rollback()
        finally:
            cur.close()
            conn.close()
    
    return redirect(url_for('clientes'))

@app.route('/modificar_cliente/<int:id>', methods=['POST'])
def modificar_cliente(id):
    # Datos básicos
    nombre = request.form['nombre'].upper()
    telefono = request.form['telefono']
    correo = request.form['correo']
    tipo = request.form['tipo']
    direccion = request.form['direccion'].upper()
    
    # Datos fiscales opcionales
    rfc = request.form.get('rfc', '').upper() or None
    razon_social = request.form.get('razon_social', '').upper() or None
    uso_cfdi = request.form.get('uso_cfdi') or None
    regimen_fiscal = request.form.get('regimen_fiscal') or None
    correo_facturacion = request.form.get('correo_facturacion') or None

    conn = get_connection()
    if conn:
        cur = conn.cursor()
        
        try:
            # Intentar actualizar con todos los campos incluyendo datos fiscales
            cur.execute("""
                UPDATE clientes 
                SET nombre_cliente=?, telefono=?, correo=?, tipo_cliente=?, direccion=?,
                    rfc=?, razon_social=?, uso_cfdi=?, regimen_fiscal=?, correo_facturacion=?
                WHERE id_cliente=?;
            """, (nombre, telefono, correo, tipo, direccion, 
                  rfc, razon_social, uso_cfdi, regimen_fiscal, correo_facturacion, id))
            conn.commit()
            print(f"✓ Cliente #{id} actualizado correctamente con datos fiscales")
        except Exception as e:
            # Si falla (porque las columnas no existen), actualizar solo campos básicos
            print(f"⚠️ Error actualizando con datos fiscales: {e}")
            try:
                cur.execute("""
                    UPDATE clientes 
                    SET nombre_cliente=?, telefono=?, correo=?, tipo_cliente=?, direccion=?
                    WHERE id_cliente=?;
                """, (nombre, telefono, correo, tipo, direccion, id))
                conn.commit()
                print(f"✓ Cliente #{id} actualizado (solo datos básicos)")
            except Exception as e2:
                print(f"❌ Error al actualizar cliente: {e2}")
                conn.rollback()
        
        cur.close()
        conn.close()
    
    return redirect(url_for('clientes'))

# ==================== CATÁLOGO ====================
@app.route('/catalogo')
def catalogo():
    conn = get_connection()
    combinaciones = []
    if conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT 
                c.id_combinacion,
                c.nombre_guardado,
                c.esquema,
                c.fecha_creacion,
                mb.nombre_modelo,
                mb.tipo,
                cp.codigo_hex AS hex_principal,
                ca.codigo_hex AS hex_asa,
                cs.codigo_hex AS hex_secundario,
                c.diseno_json
            FROM combinaciones c
            JOIN modelos_bolsas mb ON c.id_modelo = mb.id_modelo
            LEFT JOIN colores cp ON c.id_color_principal = cp.id_color
            LEFT JOIN colores ca ON c.id_color_asa = ca.id_color
            LEFT JOIN colores cs ON c.id_color_secundario = cs.id_color
            ORDER BY c.fecha_creacion DESC;
        """)
        combinaciones = cur.fetchall()
        cur.close()
        conn.close()
    return render_template('catalogo.html', combinaciones=combinaciones)

@app.route('/eliminar_combinacion/<int:id>', methods=['POST'])
def eliminar_combinacion(id):
    """Elimina un diseño del catálogo"""
    conn = get_connection()
    if conn:
        try:
            cur = conn.cursor()
            
            # Verificar si la combinación está siendo usada en cotizaciones
            cur.execute("""
                SELECT COUNT(*) FROM detalle_cotizacion 
                WHERE id_material = ?
            """, (id,))
            en_cotizaciones = cur.fetchone()[0]
            
            # Verificar si la combinación está siendo usada en pedidos
            cur.execute("""
                SELECT COUNT(*) FROM detalle_pedido 
                WHERE id_producto = ?
            """, (id,))
            en_pedidos = cur.fetchone()[0]
            
            if en_cotizaciones > 0 or en_pedidos > 0:
                print(f"⚠️ No se puede eliminar: diseño usado en {en_cotizaciones} cotizaciones y {en_pedidos} pedidos")
            else:
                # Eliminar la combinación
                cur.execute("DELETE FROM combinaciones WHERE id_combinacion = ?", (id,))
                conn.commit()
                print(f"✓ Diseño #{id} eliminado correctamente")
            
            cur.close()
        except Exception as e:
            print(f"❌ Error al eliminar combinación: {e}")
            if conn:
                conn.rollback()
        finally:
            if conn:
                conn.close()
    
    return redirect(url_for('catalogo'))

# ==================== DISEÑO COLOR ====================
@app.route('/diseno_color', methods=['GET', 'POST'])
def diseno_color():
    if request.method == 'POST':
        esquema = request.form['esquema_color']
        id_modelo = request.form['modelo_bolsa']
        nombre = request.form['nombre_combinacion']

        color_principal = request.form.get('color_principal')
        color_secundario = request.form.get('color_secundario')
        color_asa = request.form.get('color_asa')
        elementos_json = request.form.get('elementos_json')

        conn = get_connection()
        if conn:
            cur = conn.cursor()

            # Validar si el nombre ya existe
            cur.execute("SELECT COUNT(*) FROM combinaciones WHERE nombre_guardado = ?", (nombre,))
            existe = cur.fetchone()[0]
            
            if existe > 0:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': 'Ya existe una combinación con ese nombre'}), 400

            def obtener_id_color(hex_code):
                if not hex_code:
                    return None
                cur.execute("""
                    INSERT OR IGNORE INTO colores (nombre_color, codigo_hex)
                    VALUES (?, ?)
                """, (f"Color_{hex_code}", hex_code))
                conn.commit()
                cur.execute("SELECT id_color FROM colores WHERE codigo_hex = ?", (hex_code,))
                result = cur.fetchone()
                return result[0] if result else None

            id_color_principal = obtener_id_color(color_principal)
            id_color_secundario = obtener_id_color(color_secundario)
            id_color_asa = obtener_id_color(color_asa)

            cur.execute("""
                INSERT INTO combinaciones (
                    id_modelo, esquema, id_color_principal, 
                    id_color_secundario, id_color_asa, nombre_guardado, diseno_json
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (id_modelo, esquema, id_color_principal, id_color_secundario, 
                  id_color_asa, nombre, elementos_json))
            conn.commit()
            cur.close()
            conn.close()

        return redirect(url_for('catalogo'))

    conn = get_connection()
    modelos, paletas, combinaciones = [], [], []
    if conn:
        cur = conn.cursor()
        cur.execute("SELECT id_modelo, nombre_modelo, tipo FROM modelos_bolsas ORDER BY id_modelo;")
        modelos = cur.fetchall()
        cur.execute("SELECT id_paleta, nombre, esquema FROM paletas_colores ORDER BY id_paleta;")
        paletas = cur.fetchall()
        cur.execute("""
            SELECT c.id_combinacion, c.nombre_guardado, c.esquema, c.fecha_creacion,
                   mb.nombre_modelo, mb.tipo
            FROM combinaciones c
            JOIN modelos_bolsas mb ON c.id_modelo = mb.id_modelo
            ORDER BY c.fecha_creacion DESC;
        """)
        combinaciones = cur.fetchall()
        cur.close()
        conn.close()

    return render_template('diseno_color.html', modelos=modelos, paletas=paletas, 
                         combinaciones=combinaciones)

# ==================== INVENTARIO ====================
@app.route('/inventario')
def inventario():
    inventario_data = InventoryManager.obtener_inventario_completo()
    materiales = InventoryManager.obtener_materiales()
    return render_template('inventario.html', inventario=inventario_data, materiales=materiales)

@app.route('/agregar_material', methods=['POST'])
def agregar_material():
    nombre = request.form['nombre_material']
    tipo = request.form['tipo']
    unidad_medida = request.form['unidad_medida']
    costo_unitario = float(request.form['costo_unitario'])
    descripcion = request.form.get('descripcion')
    
    resultado = InventoryManager.agregar_material(
        nombre, tipo, unidad_medida, costo_unitario, descripcion
    )
    
    if not resultado['success']:
        print(f"Error al agregar material: {resultado['error']}")
    
    return redirect(url_for('inventario'))

@app.route('/actualizar_stock', methods=['POST'])
def actualizar_stock():
    id_material = int(request.form['id_material'])
    cantidad = float(request.form['cantidad'])
    
    resultado = InventoryManager.actualizar_stock(id_material, cantidad)
    
    if not resultado['success']:
        print(f"Error al actualizar stock: {resultado['error']}")
    
    return redirect(url_for('inventario'))

@app.route('/api/verificar_material', methods=['POST'])
def api_verificar_material():
    data = request.json
    id_material = data.get('id_material')
    cantidad = data.get('cantidad')
    
    disponible = InventoryManager.verificar_disponibilidad(id_material, cantidad)
    
    return jsonify({'disponible': disponible})

@app.route('/api/materiales_bajo_stock')
def api_materiales_bajo_stock():
    umbral = request.args.get('umbral', 100, type=int)
    materiales = InventoryManager.obtener_materiales_bajo_stock(umbral)
    return jsonify(materiales)

@app.route('/eliminar_material/<int:id>', methods=['POST'])
def eliminar_material(id):
    resultado = InventoryManager.eliminar_material(id)
    if not resultado['success']:
        print(f"Error al eliminar material: {resultado['error']}")
    return redirect(url_for('inventario'))

@app.route('/exportar_inventario_excel')
def exportar_inventario_excel():
    """Exporta el inventario a Excel con logo y título"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.drawing.image import Image as XLImage
        from io import BytesIO
        
        # Obtener datos del inventario
        inventario = InventoryManager.obtener_inventario_completo()
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        
        # Intentar agregar logo
        logo_path = os.path.join(app.static_folder, 'images', 'logo_.png')
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.width = 80
                img.height = 80
                ws.add_image(img, 'A1')
            except:
                pass
        
        # Título del sistema (centrado en las columnas)
        ws.merge_cells('B1:F1')
        titulo_cell = ws['B1']
        titulo_cell.value = 'Sistema Integral de Gestión para Confeccionistas de Bolsas'
        titulo_cell.font = Font(size=14, bold=True, color='FF1493')
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Subtítulo ChromaBags
        ws.merge_cells('B2:F2')
        subtitulo_cell = ws['B2']
        subtitulo_cell.value = 'CHROMABAGS'
        subtitulo_cell.font = Font(size=18, bold=True, color='FF69B4')
        subtitulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Título del documento
        ws.merge_cells('A4:F4')
        doc_title = ws['A4']
        doc_title.value = 'REPORTE DE INVENTARIO'
        doc_title.font = Font(size=16, bold=True)
        doc_title.alignment = Alignment(horizontal='center', vertical='center')
        
        # Fecha de generación
        ws.merge_cells('A5:F5')
        fecha_cell = ws['A5']
        fecha_cell.value = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        fecha_cell.font = Font(size=10, italic=True)
        fecha_cell.alignment = Alignment(horizontal='center')
        
        # Encabezados de la tabla (fila 7)
        headers = ['Material', 'Tipo', 'Stock Actual', 'Unidad', 'Costo Unit.', 'Valor Total']
        header_fill = PatternFill(start_color='FF69B4', end_color='FF69B4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=7, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos del inventario
        row_num = 8
        total_valor = 0
        
        for item in inventario:
            ws.cell(row=row_num, column=1, value=item['nombre_material']).border = border
            ws.cell(row=row_num, column=2, value=item['tipo']).border = border
            ws.cell(row=row_num, column=3, value=item['stock_actual']).border = border
            ws.cell(row=row_num, column=4, value=item['unidad_medida']).border = border
            ws.cell(row=row_num, column=5, value=f"${item['costo_unitario']:.2f}").border = border
            
            valor_total_item = item['stock_actual'] * item['costo_unitario']
            total_valor += valor_total_item
            ws.cell(row=row_num, column=6, value=f"${valor_total_item:.2f}").border = border
            
            row_num += 1
        
        # Fila de totales
        ws.cell(row=row_num, column=5, value="TOTAL:").font = Font(bold=True, size=12)
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='right')
        total_cell = ws.cell(row=row_num, column=6, value=f"${total_valor:.2f}")
        total_cell.font = Font(bold=True, size=12, color='FF1493')
        total_cell.fill = PatternFill(start_color='FFE4E1', end_color='FFE4E1', fill_type='solid')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'inventario_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        print(f"Error al exportar inventario: {e}")
        return f"Error al exportar: {str(e)}", 500
    
@app.route('/api/material/<int:id>')
def api_material(id):
    material = InventoryManager.obtener_material_por_id(id)
    return jsonify(material)

@app.route('/modificar_material/<int:id>', methods=['POST'])
def modificar_material(id):
    nombre = request.form['nombre_material']
    tipo = request.form['tipo']
    unidad = request.form['unidad_medida']
    costo = float(request.form['costo_unitario'])
    descripcion = request.form.get('descripcion')

    resultado = InventoryManager.modificar_material(id, nombre, tipo, unidad, costo, descripcion)

    if not resultado['success']:
        print("Error modificando material:", resultado['error'])

    return redirect(url_for('inventario'))


# ==================== COTIZACIÓN ====================
@app.route('/cotizacion')
def cotizacion():
    cotizaciones = QuotationManager.obtener_cotizaciones()
    
    conn = get_connection()
    clientes_data = []
    if conn:
        cur = conn.cursor()
        cur.execute("SELECT id_cliente, nombre_cliente FROM clientes ORDER BY nombre_cliente")
        clientes_data = cur.fetchall()
        cur.close()
        conn.close()
    
    productos = QuotationManager.obtener_productos_disponibles()
    
    return render_template('cotizacion.html', 
                         cotizaciones=cotizaciones, 
                         clientes=clientes_data,
                         productos=productos)

@app.route('/generar_cotizacion', methods=['POST'])
def generar_cotizacion():
    id_cliente = int(request.form['id_cliente'])
    
    # Obtener conexión para consultar tipo de modelo
    conn = get_connection()
    cur = conn.cursor()
    
    productos = []
    index = 1
    while f'productos[{index}][id_combinacion]' in request.form:
        id_combinacion = request.form.get(f'productos[{index}][id_combinacion]')
        cantidad = request.form.get(f'productos[{index}][cantidad]')
        
        if id_combinacion and cantidad:
            # Obtener el tipo de modelo para calcular el precio correcto
            cur.execute("""
                SELECT mb.tipo 
                FROM combinaciones c
                JOIN modelos_bolsas mb ON c.id_modelo = mb.id_modelo
                WHERE c.id_combinacion = ?
            """, (int(id_combinacion),))
            
            resultado_tipo = cur.fetchone()
            
            # Calcular precio según tipo de modelo
            if resultado_tipo:
                tipo_modelo = resultado_tipo[0]
                if tipo_modelo == 'combinado':
                    precio = 220
                elif tipo_modelo == 'especial':
                    precio = 250
                else:  # armonico
                    precio = 180
            else:
                precio = 180  # precio por defecto
            
            productos.append({
                'id_combinacion': int(id_combinacion),
                'cantidad': int(cantidad),
                'precio_unitario': precio
            })
        
        index += 1
    
    cur.close()
    conn.close()
    
    if not productos:
        return redirect(url_for('cotizacion'))
    
    resultado = QuotationManager.crear_cotizacion(id_cliente, productos)
    
    if not resultado['success']:
        print(f"Error al crear cotización: {resultado['error']}")
    
    return redirect(url_for('cotizacion'))

@app.route('/api/cotizacion/<int:id>')
def api_cotizacion_detalle(id):
    cotizacion = QuotationManager.obtener_cotizacion_detalle(id)
    return jsonify(cotizacion)

@app.route('/api/cotizacion/<int:id>/estado', methods=['PUT'])
def api_actualizar_estado_cotizacion(id):
    data = request.json
    nuevo_estado = data.get('estado')
    
    resultado = QuotationManager.actualizar_estado_cotizacion(id, nuevo_estado)
    return jsonify(resultado)

@app.route('/eliminar_cotizacion/<int:id>', methods=['POST'])
def eliminar_cotizacion(id):
    resultado = QuotationManager.eliminar_cotizacion(id)
    if not resultado['success']:
        print(f"Error al eliminar cotización: {resultado['error']}")
    return redirect(url_for('cotizacion'))

@app.route('/duplicar_cotizacion/<int:id>', methods=['POST'])
def duplicar_cotizacion(id):
    resultado = QuotationManager.duplicar_cotizacion(id)
    if not resultado['success']:
        print(f"Error al duplicar cotización: {resultado['error']}")
    return redirect(url_for('cotizacion'))

@app.route('/api/productos_cotizacion')
def api_productos_cotizacion():
    productos = QuotationManager.obtener_productos_disponibles()
    return jsonify(productos)

@app.route('/exportar_cotizacion/<int:id>')
def exportar_cotizacion(id):
    """Exporta una cotización individual a PDF"""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from io import BytesIO
    
    cotizacion = QuotationManager.obtener_cotizacion_detalle(id)
    if not cotizacion:
        return "Cotización no encontrada", 404
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Dibujar encabezado con logo y título
    y = dibujar_encabezado_pdf(c, width, height, f"COTIZACIÓN #{cotizacion['id_cotizacion']}")
    
    # Info cotización
    c.setFont("Helvetica", 12)
    y -= 20
    c.drawString(40, y, f"Cliente: {cotizacion['nombre_cliente']}")
    y -= 18
    c.drawString(40, y, f"Fecha de emisión: {cotizacion['fecha_emision'][:10]}")
    y -= 18
    c.drawString(40, y, f"Estado: {cotizacion['estado'].upper()}")
    y -= 35
    
    # Productos
    c.setFont("Helvetica-Bold", 13)
    c.drawString(40, y, "Detalle de Productos:")
    y -= 25
    
    c.setFont("Helvetica", 10)
    for prod in cotizacion['productos']:
        # Obtener nombre del producto
        conn = get_connection()
        if conn:
            cur = conn.cursor()
            cur.execute("SELECT nombre_guardado FROM combinaciones WHERE id_combinacion = ?", (prod['id_producto'],))
            nombre = cur.fetchone()
            nombre_producto = nombre[0] if nombre else f"Producto #{prod['id_producto']}"
            cur.close()
            conn.close()
        else:
            nombre_producto = f"Producto #{prod['id_producto']}"
        
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, y, f"• {nombre_producto}")
        y -= 16
        c.setFont("Helvetica", 10)
        c.drawString(70, y, f"Cantidad: {prod['cantidad']} unidades")
        y -= 14
        c.drawString(70, y, f"Precio unitario: ${prod['precio_unitario']:.2f}")
        y -= 14
        c.drawString(70, y, f"Subtotal: ${prod['subtotal']:.2f}")
        y -= 22
        
        if y < 100:  # Nueva página si es necesario
            c.showPage()
            y = dibujar_encabezado_pdf(c, width, height, f"COTIZACIÓN #{cotizacion['id_cotizacion']} (continuación)")
            y -= 20
    
    # Total
    y -= 15
    c.setStrokeColorRGB(0.7, 0.3, 0.5)
    c.line(40, y, width - 40, y)
    y -= 25
    c.setFont("Helvetica-Bold", 15)
    c.drawString(40, y, f"TOTAL: ${cotizacion['total_estimado']:.2f}")
    
    c.save()
    buffer.seek(0)
    
    return send_file(buffer, 
                     mimetype='application/pdf',
                     as_attachment=True,
                     download_name=f'cotizacion_{id}.pdf')

@app.route('/exportar_todas_cotizaciones')
def exportar_todas_cotizaciones():
    """Exporta todas las cotizaciones a PDF"""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from io import BytesIO
    
    cotizaciones = QuotationManager.obtener_cotizaciones()
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Dibujar encabezado con logo y título
    y = dibujar_encabezado_pdf(c, width, height, "HISTORIAL DE COTIZACIONES")
    
    # Fecha de generación
    c.setFont("Helvetica", 10)
    y -= 10
    c.drawString(40, y, f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    y -= 10
    c.drawString(40, y, f"Total de cotizaciones: {len(cotizaciones)}")
    y -= 30
    
    c.setFont("Helvetica", 11)
    
    for cot in cotizaciones:
        if y < 100:
            c.showPage()
            y = dibujar_encabezado_pdf(c, width, height, "HISTORIAL DE COTIZACIONES (continuación)")
            y -= 20
        
        c.setFont("Helvetica-Bold", 12)
        c.drawString(40, y, f"Cotización #{cot['id_cotizacion']}")
        y -= 18
        
        c.setFont("Helvetica", 10)
        c.drawString(60, y, f"Cliente: {cot['nombre_cliente']}")
        y -= 15
        c.drawString(60, y, f"Fecha: {cot['fecha_emision'][:10]}")
        y -= 15
        c.drawString(60, y, f"Productos: {cot['cantidad_productos']} ({cot['cantidad_total']} unidades)")
        y -= 15
        c.drawString(60, y, f"Total: ${cot['total_estimado']:.2f}")
        y -= 15
        c.drawString(60, y, f"Estado: {cot['estado'].upper()}")
        y -= 25
        
        # Línea separadora
        c.setStrokeColorRGB(0.8, 0.8, 0.8)
        c.line(40, y, width - 40, y)
        y -= 10
    
    c.save()
    buffer.seek(0)
    
    return send_file(buffer, 
                     mimetype='application/pdf',
                     as_attachment=True,
                     download_name='historial_cotizaciones.pdf')

# ==================== PEDIDOS ====================
@app.route('/pedidos')
def pedidos():
    conn = get_connection()
    clientes_data = []
    productos_data = []
    
    if conn:
        cur = conn.cursor()
        
        # Clientes
        cur.execute("SELECT id_cliente, nombre_cliente FROM clientes ORDER BY nombre_cliente")
        clientes_data = cur.fetchall()
        
        # Productos desde el catálogo (combinaciones)
        cur.execute("""
            SELECT 
                c.id_combinacion,
                c.nombre_guardado,
                mb.nombre_modelo,
                mb.tipo
            FROM combinaciones c
            JOIN modelos_bolsas mb ON c.id_modelo = mb.id_modelo
            ORDER BY c.nombre_guardado
        """)
        productos_data = cur.fetchall()
        
        cur.close()
        conn.close()
    
    pedidos_data = OrdersManager.obtener_pedidos()
    
    return render_template('pedidos.html',
                         clientes=clientes_data,
                         productos=productos_data,
                         pedidos=pedidos_data)

@app.route('/guardar_pedido', methods=['POST'])
def guardar_pedido():
    id_cliente = request.form['id_cliente']
    id_combinacion = request.form['id_combinacion']  # Ahora es id_combinacion
    cantidad = int(request.form['cantidad'])
    fecha_entrega = request.form['fecha_entrega']
    estado = request.form['estado']
    
    resultado = OrdersManager.crear_pedido(
        id_cliente, id_combinacion, cantidad, fecha_entrega, estado
    )
    
    if not resultado['success']:
        print(f"Error al guardar pedido: {resultado['error']}")
    
    return redirect(url_for('pedidos'))

@app.route('/actualizar_pedido/<int:id_pedido>', methods=['POST'])
def actualizar_pedido(id_pedido):
    """Actualiza fecha de entrega y estado del pedido desde la tabla"""
    fecha_entrega = request.form['fecha_entrega']
    estado = request.form['estado']
    
    resultado = OrdersManager.actualizar_pedido(id_pedido, fecha_entrega, estado)
    
    if not resultado['success']:
        print(f"Error al actualizar pedido: {resultado['error']}")
    
    return redirect(url_for('pedidos'))

@app.route('/editar_pedido/<int:id_pedido>', methods=['GET', 'POST'])
def editar_pedido(id_pedido):
    if request.method == 'POST':
        fecha_entrega = request.form['fecha_entrega']
        estado = request.form['estado']
        
        resultado = OrdersManager.actualizar_pedido(id_pedido, fecha_entrega, estado)
        
        if not resultado['success']:
            print(f"Error al actualizar pedido: {resultado['error']}")
        
        return redirect(url_for('pedidos'))
    
    pedido = OrdersManager.obtener_pedido(id_pedido)
    return render_template('editar_pedido.html', pedido=pedido)

@app.route('/eliminar_pedido/<int:id_pedido>')
def eliminar_pedido(id_pedido):
    resultado = OrdersManager.eliminar_pedido(id_pedido)
    
    if not resultado['success']:
        print(f"Error al eliminar pedido: {resultado['error']}")
    
    return redirect(url_for('pedidos'))

# ==================== PAGOS ====================
@app.route('/pagos')
def pagos():
    """Muestra pedidos finalizados con información de pagos"""
    conn = get_connection()
    pagos_data = []
    historial_data = []
    
    if conn:
        cur = conn.cursor()
        
        # Obtenemos pedidos finalizados y calculamos el total pagado
        cur.execute("""
            SELECT 
                p.id_pedido,
                c.nombre_cliente,
                p.total,
                COALESCE(SUM(pg.monto), 0) as total_pagado
            FROM pedidos p
            JOIN clientes c ON p.id_cliente = c.id_cliente
            LEFT JOIN pagos pg ON p.id_pedido = pg.id_pedido
            WHERE p.estado IN ('finalizado', 'entregado')
            GROUP BY p.id_pedido, c.nombre_cliente, p.total
            ORDER BY p.fecha_pedido DESC
        """)
        
        rows = cur.fetchall()
        for row in rows:
            id_pedido = row[0]
            nombre_cliente = row[1]
            total = row[2]
            total_pagado = row[3]
            
            # Calcular lo que falta por pagar
            falta_pagar = max(0, total - total_pagado)
            
            # Determinar estado de pago
            if total_pagado >= total:
                estado = 'pagado'
            else:
                estado = 'pendiente'
            
            pagos_data.append({
                'id_pedido': id_pedido,
                'nombre_cliente': nombre_cliente,
                'total': total,
                'total_pagado': total_pagado,
                'falta_pagar': falta_pagar,
                'estado': estado
            })
        
        # Obtener historial de todos los pagos
        cur.execute("""
            SELECT 
                pg.id_pago,
                pg.id_pedido,
                c.nombre_cliente,
                pg.monto,
                pg.metodo,
                pg.fecha_pago
            FROM pagos pg
            JOIN pedidos p ON pg.id_pedido = p.id_pedido
            JOIN clientes c ON p.id_cliente = c.id_cliente
            ORDER BY pg.fecha_pago DESC
            LIMIT 50
        """)
        
        rows = cur.fetchall()
        for row in rows:
            historial_data.append({
                'id_pago': row[0],
                'id_pedido': row[1],
                'nombre_cliente': row[2],
                'monto': row[3],
                'metodo': row[4],
                'fecha_pago': row[5]
            })
        
        cur.close()
        conn.close()
    
    return render_template('pagos.html', pagos=pagos_data, historial=historial_data)

@app.route('/registrar_pago/<int:id_pedido>', methods=['POST'])
def registrar_pago(id_pedido):
    """Registra un pago para un pedido"""
    metodo = request.form['metodo']
    monto = float(request.form['monto'])
    
    conn = get_connection()
    if conn:
        cur = conn.cursor()
        
        # Obtener el total del pedido y lo ya pagado
        cur.execute("""
            SELECT p.total, COALESCE(SUM(pg.monto), 0) as total_pagado
            FROM pedidos p
            LEFT JOIN pagos pg ON p.id_pedido = pg.id_pedido
            WHERE p.id_pedido = ?
            GROUP BY p.total
        """, (id_pedido,))
        
        row = cur.fetchone()
        if row:
            total_pedido = row[0]
            total_pagado = row[1]
            falta_pagar = total_pedido - total_pagado
            
            # Validar que el monto no exceda lo que falta
            if monto > falta_pagar:
                monto = falta_pagar
            
            # Insertar el pago
            cur.execute("""
                INSERT INTO pagos (id_pedido, monto, metodo)
                VALUES (?, ?, ?)
            """, (id_pedido, monto, metodo))
            
            conn.commit()
        
        cur.close()
        conn.close()
    
    return redirect(url_for('pagos'))

# ==================== FACTURACIÓN ====================
@app.route('/facturacion')
def facturacion():
    """Muestra pedidos completamente pagados listos para facturar"""
    conn = get_connection()
    pedidos_pagados = []
    
    if conn:
        cur = conn.cursor()
        
        # Obtener pedidos finalizados o entregados
        cur.execute("""
            SELECT 
                p.id_pedido,
                c.nombre_cliente,
                p.total,
                p.estado,
                p.fecha_pedido
            FROM pedidos p
            JOIN clientes c ON p.id_cliente = c.id_cliente
            WHERE p.estado IN ('finalizado', 'entregado')
        """)
        
        pedidos = cur.fetchall()
        
        # Verificar cuáles están completamente pagados
        for pedido in pedidos:
            id_pedido = pedido[0]
            total_pedido = pedido[2]
            
            # Calcular total pagado
            cur.execute("""
                SELECT COALESCE(SUM(monto), 0)
                FROM pagos
                WHERE id_pedido = ?
            """, (id_pedido,))
            
            total_pagado = cur.fetchone()[0]
            
            # Solo incluir si está completamente pagado
            if total_pagado >= total_pedido:
                pedidos_pagados.append({
                    'id_pedido': pedido[0],
                    'nombre_cliente': pedido[1],
                    'total': pedido[2],
                    'estado': pedido[3],
                    'fecha_pedido': pedido[4]
                })
        
        cur.close()
        conn.close()
    
    return render_template('facturacion.html', pedidos_pagados=pedidos_pagados)

@app.route('/generar_factura/<int:id_pedido>', methods=['POST'])
def generar_factura(id_pedido):
    """Genera una factura en PDF para un pedido"""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from io import BytesIO
    
    conn = get_connection()
    if not conn:
        return "Error de conexión", 500
    
    try:
        cur = conn.cursor()
        
        # Obtener información del pedido y cliente con datos fiscales
        cur.execute("""
            SELECT 
                p.id_pedido,
                c.nombre_cliente,
                c.correo,
                c.telefono,
                c.direccion,
                p.fecha_pedido,
                p.total,
                dp.cantidad,
                comb.nombre_guardado,
                c.rfc,
                c.razon_social,
                c.uso_cfdi,
                c.regimen_fiscal,
                c.correo_facturacion,
                dp.precio_unitario
            FROM pedidos p
            JOIN clientes c ON p.id_cliente = c.id_cliente
            JOIN detalle_pedido dp ON p.id_pedido = dp.id_pedido
            LEFT JOIN combinaciones comb ON dp.id_producto = comb.id_combinacion
            WHERE p.id_pedido = ?
        """, (id_pedido,))
        
        pedido_info = cur.fetchone()
        
        if not pedido_info:
            cur.close()
            conn.close()
            return "Pedido no encontrado", 404
        
        # Calcular subtotal e IVA
        total_con_iva = pedido_info[6]
        subtotal = total_con_iva / 1.16  # Calcular subtotal sin IVA
        iva = total_con_iva - subtotal  # IVA es la diferencia
        precio_unitario_sin_iva = pedido_info[14] / 1.16 if pedido_info[14] else 0
        
        # Crear PDF
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # Dibujar encabezado con logo y título
        y = dibujar_encabezado_pdf(c, width, height, f"FACTURA #{pedido_info[0]}")
        
        # Información de la factura
        c.setFont("Helvetica", 11)
        y -= 10
        c.drawString(40, y, f"Fecha de emisión: {pedido_info[5][:10] if pedido_info[5] else datetime.now().strftime('%d/%m/%Y')}")
        y -= 35
        
        # Datos del cliente
        c.setFont("Helvetica-Bold", 13)
        c.drawString(40, y, "Datos del Cliente:")
        y -= 22
        
        c.setFont("Helvetica", 10)
        c.drawString(40, y, f"Nombre: {pedido_info[1]}")
        y -= 15
        if pedido_info[10]:  # razon_social
            c.drawString(40, y, f"Razón Social: {pedido_info[10]}")
            y -= 15
        if pedido_info[9]:  # rfc
            c.drawString(40, y, f"RFC: {pedido_info[9]}")
            y -= 15
        if pedido_info[12]:  # regimen_fiscal
            c.drawString(40, y, f"Régimen Fiscal: {pedido_info[12]}")
            y -= 15
        c.drawString(40, y, f"Teléfono: {pedido_info[3] or 'N/A'}")
        y -= 15
        c.drawString(40, y, f"Dirección: {pedido_info[4] or 'N/A'}")
        y -= 15
        if pedido_info[13]:  # correo_facturacion
            c.drawString(40, y, f"Email Facturación: {pedido_info[13]}")
            y -= 15
        if pedido_info[11]:  # uso_cfdi
            c.drawString(40, y, f"Uso CFDI: {pedido_info[11]}")
            y -= 15
        y -= 20
        
        # Detalles del pedido
        c.setFont("Helvetica-Bold", 13)
        c.drawString(40, y, "Detalles del Pedido:")
        y -= 25
        
        # Encabezados de tabla
        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, y, "Producto")
        c.drawString(280, y, "Cantidad")
        c.drawString(350, y, "P. Unitario")
        c.drawString(450, y, "Subtotal")
        y -= 5
        c.setStrokeColorRGB(0.7, 0.3, 0.5)
        c.line(40, y, width - 40, y)
        y -= 20
        
        # Productos
        c.setFont("Helvetica", 10)
        c.setFillColorRGB(0, 0, 0)
        c.drawString(40, y, pedido_info[8] or "Producto sin nombre")
        c.drawString(295, y, str(pedido_info[7]))
        c.drawString(350, y, f"${precio_unitario_sin_iva:.2f}")
        subtotal_producto = precio_unitario_sin_iva * pedido_info[7]
        c.drawString(450, y, f"${subtotal_producto:.2f}")
        y -= 35
        
        # Desglose de totales
        c.setStrokeColorRGB(0.7, 0.3, 0.5)
        c.setLineWidth(1)
        c.line(280, y, width - 40, y)
        y -= 25
        
        # Subtotal
        c.setFont("Helvetica", 12)
        c.drawString(320, y, "Subtotal:")
        c.drawString(450, y, f"${subtotal:.2f}")
        y -= 20
        
        # IVA
        c.drawString(320, y, "IVA (16%):")
        c.drawString(450, y, f"${iva:.2f}")
        y -= 25
        
        # Total
        c.setStrokeColorRGB(0.7, 0.3, 0.5)
        c.setLineWidth(2)
        c.line(280, y, width - 40, y)
        y -= 25
        c.setFont("Helvetica-Bold", 15)
        c.setFillColorRGB(0.7, 0.3, 0.5)
        c.drawString(320, y, "TOTAL:")
        c.drawString(450, y, f"${total_con_iva:.2f}")
        
        # Nota sobre IVA
        y -= 40
        c.setFont("Helvetica-Oblique", 9)
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.drawString(40, y, "* Los precios incluyen IVA del 16% según la legislación vigente")
        
        # Pie de página
        c.setFont("Helvetica-Oblique", 9)
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.drawString(40, 60, "Gracias por su compra")
        c.setFont("Helvetica", 8)
        c.drawString(40, 45, "ChromaBags - Sistema Integral de Gestión para Confeccionistas de Bolsas")
        c.drawString(40, 30, "www.chromabags.com | contacto@chromabags.com")
        
        c.save()
        buffer.seek(0)
        
        cur.close()
        conn.close()
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'factura_{id_pedido}.pdf'
        )
        
    except Exception as e:
        print(f"Error generando factura: {e}")
        if conn:
            conn.close()
        return f"Error al generar factura: {str(e)}", 500

# ==================== REPORTES ====================
@app.route('/reportes')
def reportes():
    datos = OrdersManager.obtener_pedidos_por_estado()
    estadisticas = OrdersManager.obtener_estadisticas_dashboard()
    
    # Valores por defecto si no hay estadísticas
    if not estadisticas:
        estadisticas = {
            'total_pedidos': 0,
            'por_estado': {},
            'ingresos_totales': 0,
            'pedidos_mes': 0,
            'ingresos_mes': 0,
            'clientes_activos': 0,
            'productos_top': [],
            'ventas_mensuales': []
        }
    
    return render_template('reportes.html',
                         por_entregar=datos['por_entregar'],
                         entregados=datos['entregados'],
                         vencidos=datos['vencidos'],
                         todos=datos['todos'],
                         estadisticas=estadisticas)

# ==================== RESPALDO ====================
@app.route('/respaldo')
def respaldo():
    respaldos = BackupManager.obtener_respaldos()
    return render_template('respaldo.html', respaldos=respaldos)

@app.route('/descargar_respaldo')
def descargar_respaldo():
    resultado = BackupManager.crear_respaldo()
    
    if not resultado['success']:
        return resultado['error'], 500
    
    return send_file(
        resultado['ruta'],
        as_attachment=True,
        download_name=resultado['nombre']
    )

@app.route('/descargar_respaldo/<nombre_archivo>')
def descargar_respaldo_existente(nombre_archivo):
    """Descarga un respaldo existente"""
    import os
    
    # Ruta del directorio de respaldos
    directorio_respaldos = os.path.join(os.path.dirname(__file__), 'respaldos')
    ruta_archivo = os.path.join(directorio_respaldos, nombre_archivo)
    
    # Validar que el archivo existe y está en el directorio correcto
    if not os.path.exists(ruta_archivo):
        return "Archivo no encontrado", 404
    
    if not ruta_archivo.startswith(directorio_respaldos):
        return "Acceso no autorizado", 403
    
    return send_file(
        ruta_archivo,
        as_attachment=True,
        download_name=nombre_archivo
    )

# ==================== APIs DE COLOR Y DISEÑO ====================
@app.route('/api/colores_paleta/<int:id_paleta>')
def api_colores_paleta(id_paleta):
    conn = get_connection()
    colores = []
    if conn:
        cur = conn.cursor()
        cur.execute("SELECT nombre_color, codigo_hex FROM colores WHERE id_paleta = ?;", (id_paleta,))
        colores = [{'nombre': row[0], 'hex': row[1]} for row in cur.fetchall()]
        cur.close()
        conn.close()
    return jsonify(colores)

@app.route('/api/generar_diseno', methods=['POST'])
def api_generar_diseno():
    data = request.json
    tipo_modelo = data.get('tipo_modelo')
    esquema = data.get('esquema')
    colores = data.get('colores', [])
    
    try:
        designer = BagDesigner(tipo_modelo)
        
        if tipo_modelo == 'simple':
            if len(colores) < 1:
                return jsonify({'error': 'Se requiere al menos 1 color'}), 400
            design = designer.create_simple_design(colores[0])
        
        elif tipo_modelo == 'combinado':
            if len(colores) < 2:
                return jsonify({'error': 'Se requieren 2 colores'}), 400
            design = designer.create_combinado_design(colores[0], colores[1])
        
        elif tipo_modelo == 'especial':
            design = designer.create_especial_design()
        
        else:
            return jsonify({'error': 'Tipo de modelo no válido'}), 400
        
        if not designer.validate_design(design, esquema):
            return jsonify({
                'warning': 'Los colores seleccionados no cumplen con el esquema de armonía',
                'design': design
            })
        
        return jsonify({
            'success': True,
            'design': design,
            'svg': designer.get_svg_representation(design)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/validar_armonia', methods=['POST'])
def api_validar_armonia():
    data = request.json
    colores = data.get('colores', [])
    esquema = data.get('esquema', 'armonico')
    
    try:
        es_valido = ColorManager.validate_harmony(colores, esquema)
        
        sugerencias = []
        if not es_valido:
            if esquema == 'complementario' and len(colores) >= 1:
                sugerencias.append({
                    'tipo': 'complementario',
                    'color': ColorManager.get_complementary(colores[0])
                })
            elif esquema == 'analogo' and len(colores) >= 1:
                sugerencias.extend([
                    {'tipo': 'analogo', 'color': c}
                    for c in ColorManager.get_analogous(colores[0])
                ])
        
        return jsonify({
            'valido': es_valido,
            'esquema': esquema,
            'sugerencias': sugerencias
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sugerir_asa', methods=['POST'])
def api_sugerir_asa():
    data = request.json
    colores_cuerpo = data.get('colores', [])
    
    try:
        color_sugerido = ColorManager.suggest_handle_color(colores_cuerpo)
        return jsonify({
            'color_sugerido': color_sugerido,
            'nombre': 'Blanco' if color_sugerido == '#FFFFFF' else 'Negro'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== OTRAS RUTAS ====================
@app.route('/ver_combinacion/<int:id>')
def ver_combinacion(id):
    conn = get_connection()
    combinacion = None
    colores = []
    
    if conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT c.*, mb.nombre_modelo, mb.tipo
            FROM combinaciones c
            JOIN modelos_bolsas mb ON c.id_modelo = mb.id_modelo
            WHERE c.id_combinacion = ?
        """, (id,))
        combinacion = cur.fetchone()
        
        if combinacion:
            color_ids = [
                combinacion[3],
                combinacion[4],
                combinacion[5],
                combinacion[6]
            ]
            
            for cid in color_ids:
                if cid:
                    cur.execute("SELECT nombre_color, codigo_hex FROM colores WHERE id_color = ?", (cid,))
                    color = cur.fetchone()
                    if color:
                        colores.append(color)
        
        cur.close()
        conn.close()
    
    return render_template('ver_combinacion.html', combinacion=combinacion, colores=colores)

if __name__ == '__main__':
    app.run(debug=True, host="127.0.0.1", port=5050)